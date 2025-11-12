"""
Results Export Module
=====================

This module provides utilities for exporting optimization results to various formats.

Supported Formats:
    - CSV files
    - Excel files (single and multi-sheet)
    - JSON files
    - PDF reports (future)

Features:
    - Automatic formatting
    - Multi-sheet Excel workbooks
    - Summary and detailed exports
    - Timestamped filenames
    - Compression options

Usage:
    from utils.exporter import ResultsExporter

    exporter = ResultsExporter()
    exporter.export_to_csv(results_df, 'results.csv')
    exporter.export_to_excel_multi_sheet({
        'Summary': summary_df,
        'Baseline': baseline_df,
        'Optimized': optimized_df
    }, 'optimization_results.xlsx')
"""

import pandas as pd
import numpy as np
from pathlib import Path
from typing import Union, Dict, Optional, List
from datetime import datetime
import json

from utils.logger import get_logger
from utils.exceptions import DataError

logger = get_logger(__name__)


# ============================================================================
# RESULTS EXPORTER CLASS
# ============================================================================

class ResultsExporter:
    """
    Export optimization results to various formats.

    This class handles exporting results with proper formatting and metadata.
    """

    def __init__(self, add_timestamp: bool = False):
        """
        Initialize ResultsExporter.

        Parameters
        ----------
        add_timestamp : bool, default=False
            Add timestamp to exported filenames
        """
        self.add_timestamp = add_timestamp

    # ========================================================================
    # CSV EXPORT
    # ========================================================================

    def export_to_csv(
        self,
        data: pd.DataFrame,
        file_path: Union[str, Path],
        **kwargs
    ) -> Path:
        """
        Export data to CSV file.

        Parameters
        ----------
        data : pd.DataFrame
            Data to export
        file_path : str or Path
            Output file path
        **kwargs
            Additional arguments for pd.to_csv

        Returns
        -------
        Path
            Path to exported file

        Examples
        --------
        >>> exporter = ResultsExporter()
        >>> exporter.export_to_csv(results_df, 'results.csv')
        """
        file_path = self._prepare_file_path(file_path)

        logger.info(f"Exporting to CSV: {file_path}")

        try:
            # Ensure directory exists
            file_path.parent.mkdir(parents=True, exist_ok=True)

            # Export
            data.to_csv(file_path, index=False, **kwargs)

            logger.info(
                f"Successfully exported to CSV",
                file_path=str(file_path),
                rows=len(data),
                columns=len(data.columns)
            )

            return file_path

        except Exception as e:
            logger.error(f"Failed to export CSV: {str(e)}")
            raise DataError(
                f"CSV export failed: {str(e)}",
                details={"file_path": str(file_path), "error": str(e)}
            )

    # ========================================================================
    # EXCEL EXPORT
    # ========================================================================

    def export_to_excel(
        self,
        data: pd.DataFrame,
        file_path: Union[str, Path],
        sheet_name: str = 'Data',
        **kwargs
    ) -> Path:
        """
        Export data to Excel file.

        Parameters
        ----------
        data : pd.DataFrame
            Data to export
        file_path : str or Path
            Output file path
        sheet_name : str, default='Data'
            Name of the Excel sheet
        **kwargs
            Additional arguments for pd.to_excel

        Returns
        -------
        Path
            Path to exported file

        Examples
        --------
        >>> exporter = ResultsExporter()
        >>> exporter.export_to_excel(results_df, 'results.xlsx')
        """
        file_path = self._prepare_file_path(file_path, extension='.xlsx')

        logger.info(f"Exporting to Excel: {file_path}")

        try:
            # Ensure directory exists
            file_path.parent.mkdir(parents=True, exist_ok=True)

            # Export
            data.to_excel(
                file_path,
                sheet_name=sheet_name,
                index=False,
                engine='openpyxl',
                **kwargs
            )

            logger.info(
                f"Successfully exported to Excel",
                file_path=str(file_path),
                rows=len(data),
                columns=len(data.columns),
                sheet_name=sheet_name
            )

            return file_path

        except Exception as e:
            logger.error(f"Failed to export Excel: {str(e)}")
            raise DataError(
                f"Excel export failed: {str(e)}",
                details={"file_path": str(file_path), "error": str(e)}
            )

    def export_to_excel_multi_sheet(
        self,
        data_dict: Dict[str, pd.DataFrame],
        file_path: Union[str, Path],
        format_sheets: bool = True
    ) -> Path:
        """
        Export multiple DataFrames to multi-sheet Excel file.

        Parameters
        ----------
        data_dict : Dict[str, pd.DataFrame]
            Dictionary mapping sheet names to DataFrames
        file_path : str or Path
            Output file path
        format_sheets : bool, default=True
            Apply formatting to sheets

        Returns
        -------
        Path
            Path to exported file

        Examples
        --------
        >>> exporter = ResultsExporter()
        >>> exporter.export_to_excel_multi_sheet({
        ...     'Summary': summary_df,
        ...     'Baseline': baseline_df,
        ...     'Optimized': optimized_df
        ... }, 'results.xlsx')
        """
        file_path = self._prepare_file_path(file_path, extension='.xlsx')

        logger.info(
            f"Exporting multi-sheet Excel: {file_path}",
            sheets=list(data_dict.keys())
        )

        try:
            # Ensure directory exists
            file_path.parent.mkdir(parents=True, exist_ok=True)

            # Create Excel writer
            with pd.ExcelWriter(
                file_path,
                engine='openpyxl',
                mode='w'
            ) as writer:
                for sheet_name, df in data_dict.items():
                    df.to_excel(
                        writer,
                        sheet_name=sheet_name,
                        index=False
                    )

                    # Apply formatting if requested
                    if format_sheets:
                        self._format_excel_sheet(
                            writer,
                            sheet_name,
                            df
                        )

            logger.info(
                f"Successfully exported multi-sheet Excel",
                file_path=str(file_path),
                sheets=len(data_dict)
            )

            return file_path

        except Exception as e:
            logger.error(f"Failed to export multi-sheet Excel: {str(e)}")
            raise DataError(
                f"Multi-sheet Excel export failed: {str(e)}",
                details={"file_path": str(file_path), "error": str(e)}
            )

    def _format_excel_sheet(
        self,
        writer: pd.ExcelWriter,
        sheet_name: str,
        df: pd.DataFrame
    ):
        """Apply formatting to Excel sheet."""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter

            workbook = writer.book
            worksheet = workbook[sheet_name]

            # Format header row
            header_fill = PatternFill(
                start_color='366092',
                end_color='366092',
                fill_type='solid'
            )
            header_font = Font(color='FFFFFF', bold=True)

            for col_num, col_name in enumerate(df.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Auto-adjust column widths
            for col_num, col_name in enumerate(df.columns, 1):
                column_letter = get_column_letter(col_num)
                max_length = max(
                    len(str(col_name)),
                    df[col_name].astype(str).str.len().max()
                )
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Format numeric columns
            for col_num, col_name in enumerate(df.columns, 1):
                if pd.api.types.is_numeric_dtype(df[col_name]):
                    column_letter = get_column_letter(col_num)
                    for row_num in range(2, len(df) + 2):
                        cell = worksheet[f'{column_letter}{row_num}']
                        # Apply number format
                        if df[col_name].dtype == 'float64':
                            cell.number_format = '#,##0.00'
                        else:
                            cell.number_format = '#,##0'

        except ImportError:
            logger.warning("openpyxl not available for formatting")
        except Exception as e:
            logger.warning(f"Excel formatting failed: {str(e)}")

    # ========================================================================
    # JSON EXPORT
    # ========================================================================

    def export_to_json(
        self,
        data: Union[pd.DataFrame, dict],
        file_path: Union[str, Path],
        orient: str = 'records',
        indent: int = 2
    ) -> Path:
        """
        Export data to JSON file.

        Parameters
        ----------
        data : pd.DataFrame or dict
            Data to export
        file_path : str or Path
            Output file path
        orient : str, default='records'
            JSON orientation for DataFrame
        indent : int, default=2
            JSON indentation

        Returns
        -------
        Path
            Path to exported file

        Examples
        --------
        >>> exporter = ResultsExporter()
        >>> exporter.export_to_json(results_df, 'results.json')
        """
        file_path = self._prepare_file_path(file_path, extension='.json')

        logger.info(f"Exporting to JSON: {file_path}")

        try:
            # Ensure directory exists
            file_path.parent.mkdir(parents=True, exist_ok=True)

            # Export based on data type
            if isinstance(data, pd.DataFrame):
                data.to_json(
                    file_path,
                    orient=orient,
                    indent=indent
                )
            elif isinstance(data, dict):
                with open(file_path, 'w') as f:
                    json.dump(data, f, indent=indent, default=str)
            else:
                raise ValueError(f"Unsupported data type: {type(data)}")

            logger.info(
                f"Successfully exported to JSON",
                file_path=str(file_path)
            )

            return file_path

        except Exception as e:
            logger.error(f"Failed to export JSON: {str(e)}")
            raise DataError(
                f"JSON export failed: {str(e)}",
                details={"file_path": str(file_path), "error": str(e)}
            )

    # ========================================================================
    # SUMMARY EXPORT
    # ========================================================================

    def export_summary_report(
        self,
        comparison: dict,
        file_path: Union[str, Path],
        format: str = 'txt'
    ) -> Path:
        """
        Export optimization summary report.

        Parameters
        ----------
        comparison : dict
            Comparison dictionary from optimizer
        file_path : str or Path
            Output file path
        format : str, default='txt'
            Output format ('txt', 'md', 'json')

        Returns
        -------
        Path
            Path to exported file

        Examples
        --------
        >>> exporter = ResultsExporter()
        >>> exporter.export_summary_report(comparison, 'summary.txt')
        """
        file_path = self._prepare_file_path(file_path)

        logger.info(f"Exporting summary report: {file_path}")

        try:
            # Ensure directory exists
            file_path.parent.mkdir(parents=True, exist_ok=True)

            if format == 'json':
                with open(file_path, 'w') as f:
                    json.dump(comparison, f, indent=2, default=str)
            else:
                # Generate text report
                report = self._generate_text_summary(comparison)
                with open(file_path, 'w') as f:
                    f.write(report)

            logger.info(f"Successfully exported summary report: {file_path}")
            return file_path

        except Exception as e:
            logger.error(f"Failed to export summary: {str(e)}")
            raise DataError(
                f"Summary export failed: {str(e)}",
                details={"file_path": str(file_path), "error": str(e)}
            )

    def _generate_text_summary(self, comparison: dict) -> str:
        """Generate text summary from comparison dictionary."""
        lines = [
            "=" * 60,
            "OPTIMIZATION SUMMARY REPORT",
            "=" * 60,
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "BASELINE PERFORMANCE",
            "-" * 60
        ]

        if 'baseline' in comparison:
            baseline = comparison['baseline']
            lines.extend([
                f"Total Cost: ${baseline.get('total_cost', 0):,.2f}",
                f"Total Investment: ${baseline.get('total_investment', 0):,.2f}",
                f"Service Level: {baseline.get('avg_service_level', 0)*100:.1f}%",
                ""
            ])

        lines.extend([
            "OPTIMIZATION RESULTS",
            "-" * 60
        ])

        for method, data in comparison.items():
            if method != 'baseline' and 'cost_reduction_pct' in data:
                lines.extend([
                    f"{method.replace('_', ' ').title()}:",
                    f"  Cost Reduction: {data['cost_reduction_pct']:.2f}%",
                    f"  Total Cost: ${data['total_cost']:,.2f}",
                    ""
                ])

        lines.extend([
            "=" * 60
        ])

        return "\n".join(lines)

    # ========================================================================
    # HELPER METHODS
    # ========================================================================

    def _prepare_file_path(
        self,
        file_path: Union[str, Path],
        extension: Optional[str] = None
    ) -> Path:
        """Prepare file path with optional timestamp and extension."""
        file_path = Path(file_path)

        # Add extension if specified and missing
        if extension and not file_path.suffix:
            file_path = file_path.with_suffix(extension)

        # Add timestamp if requested
        if self.add_timestamp:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            stem = file_path.stem
            file_path = file_path.with_name(f"{stem}_{timestamp}{file_path.suffix}")

        return file_path


# ============================================================================
# CONVENIENCE FUNCTIONS
# ============================================================================

def export_optimization_results(
    baseline_results: pd.DataFrame,
    optimized_results: Optional[pd.DataFrame],
    comparison: dict,
    output_dir: Union[str, Path],
    base_name: str = 'optimization_results'
) -> Dict[str, Path]:
    """
    Export complete optimization results to multiple files.

    Parameters
    ----------
    baseline_results : pd.DataFrame
        Baseline optimization results
    optimized_results : pd.DataFrame, optional
        Optimized results
    comparison : dict
        Comparison dictionary
    output_dir : str or Path
        Output directory
    base_name : str, default='optimization_results'
        Base name for output files

    Returns
    -------
    dict
        Dictionary mapping output type to file paths

    Examples
    --------
    >>> paths = export_optimization_results(
    ...     baseline_df,
    ...     optimized_df,
    ...     comparison,
    ...     'output/'
    ... )
    >>> print(f"Excel: {paths['excel']}")
    >>> print(f"Summary: {paths['summary']}")
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    exporter = ResultsExporter()
    paths = {}

    # Export multi-sheet Excel
    sheets = {'Baseline': baseline_results}
    if optimized_results is not None:
        sheets['Optimized'] = optimized_results

    # Create comparison DataFrame
    comparison_data = []
    for method, data in comparison.items():
        comparison_data.append({
            'Method': method.replace('_', ' ').title(),
            'Total Cost': data.get('total_cost', 0),
            'Cost Reduction %': data.get('cost_reduction_pct', 0),
        })
    sheets['Comparison'] = pd.DataFrame(comparison_data)

    paths['excel'] = exporter.export_to_excel_multi_sheet(
        sheets,
        output_dir / f'{base_name}.xlsx'
    )

    # Export CSV files
    paths['baseline_csv'] = exporter.export_to_csv(
        baseline_results,
        output_dir / f'{base_name}_baseline.csv'
    )

    if optimized_results is not None:
        paths['optimized_csv'] = exporter.export_to_csv(
            optimized_results,
            output_dir / f'{base_name}_optimized.csv'
        )

    # Export summary
    paths['summary'] = exporter.export_summary_report(
        comparison,
        output_dir / f'{base_name}_summary.txt'
    )

    logger.info(f"Exported all results to {output_dir}")

    return paths
